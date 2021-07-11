from bisect import bisect_left
from openpyxl import load_workbook
import re
import numpy as np
import pickle
from prettytable import PrettyTable
from math import log10, sqrt
from heapq import heapify, heappush, heappop
import timeit


class SearchEngine:

    def _load_documents(self):

        self.content_list = []
        self.url_list = []
        self.length_list = []
        self.topic_list = []
        self.NUMBER_DOCS = 0
        self.NUMBER_RESULTS = 20
        self.NUMBER_DOCS_CHAMPION_LIST = 40
        self.HEAP_ENABLE = True
        self.CHAMPION_LIST_ENABLE = True


        input_file_path = [
            "./input_data/IR00_3_11k News.xlsx",
            "./input_data/IR00_3_17k News.xlsx",
            "./input_data/IR00_3_20k News.xlsx"
                           ]
        for path in input_file_path:
            print(path)
            wb = load_workbook(path)
            sheet = wb.active

            i = 1
            while True:

                try:
                    i += 1
                    data_id = int(sheet.cell(i, 1).value)

                    data_content = sheet.cell(i, 2).value

                    data_topic = sheet.cell(i, 3).value

                    data_url = sheet.cell(i, 4).value
                    self.content_list.append(data_content)
                    self.topic_list.append(data_topic)
                    self.url_list.append(data_url)
                    self.NUMBER_DOCS += 1
                    if self.NUMBER_DOCS%5000 == 0:
                        print(f"Number Docs = {self.NUMBER_DOCS}")

                except Exception:
                    break


    def _load_lemmatize(self):
        file = open("./input_data/lemmatize")
        self.lemmatize = []
        for line in file.readlines():
            past, present = line.split(" ")
            self.lemmatize.append({"past": past, "present": present})

    def _load_mokassar(self):
        file = open("./input_data/mokassar")
        self.mokassar = []
        for line in file.readlines():
            plural, single = line.split(" ")
            self.mokassar.append({"plural": plural, "single": single})

    def _load_stop_normalization(self):
        file = open("./input_data/stop_normalization")
        self.stop_normalization = set()
        for word in file.readlines():
            self.stop_normalization.add(word.strip())

    def _load_input_data(self):
        self._load_documents()
        self._load_lemmatize()
        self._load_mokassar()
        self._load_stop_normalization()

    @staticmethod
    def _preprocess_remove_redundant_symbol(text):
        text = re.sub('\.', ' ', text)
        text = re.sub('،', ' ', text)
        text = re.sub(',', ' ', text)
        text = re.sub('\n', ' ', text)
        text = re.sub(':', ' ', text)
        text = re.sub('؛', ' ', text)
        text = re.sub('"', ' ', text)
        text = re.sub('“', ' ', text)
        text = re.sub('”', ' ', text)
        text = re.sub('\(', ' ', text)
        text = re.sub('\)', ' ', text)
        text = re.sub('\]', ' ', text)
        text = re.sub('\[', ' ', text)
        text = re.sub('-', ' ', text)
        text = re.sub('\*', ' ', text)
        text = re.sub('«', ' ', text)
        text = re.sub('»', ' ', text)
        text = re.sub('؟', ' ', text)
        text = re.sub('\?', ' ', text)
        text = re.sub('/', ' ', text)
        text = re.sub('!', ' ', text)
        text = re.sub('-', ' ', text)
        text = re.sub('–', ' ', text)
        text = re.sub('ـ', ' ', text)
        text = re.sub('_', ' ', text)
        text = re.sub('…', ' ', text)

        text = re.sub(u"\u064D", ' ', text)
        text = re.sub(u"\u0650", ' ', text)
        text = re.sub(u"\u064B", ' ', text)
        text = re.sub(u"\u064E", ' ', text)
        text = re.sub(u"\u064C", ' ', text)
        text = re.sub(u"\u0651", ' ', text)

        text = re.sub(r' +', ' ', text)
        return text

    @staticmethod
    def _preprocess_equalize(text):
        text = re.sub("ك", "ک", text)
        text = re.sub("اً", "ا", text)
        text = re.sub("أ", "ا", text)
        text = re.sub("ؤً", "و", text)
        text = re.sub("ة", "ه", text)
        text = re.sub("ي", "ی", text)

        text = re.sub("1", "۱", text)
        text = re.sub("2", "۲", text)
        text = re.sub("3", "۳", text)
        text = re.sub("4", "۴", text)
        text = re.sub("5", "۵", text)
        text = re.sub("6", "۶", text)
        text = re.sub("7", "۷", text)
        text = re.sub("8", "۸", text)
        text = re.sub("9", "۹", text)
        text = re.sub("0", "۰", text)
        return text

    @staticmethod
    def preprocess(text):
        text = SearchEngine._preprocess_remove_redundant_symbol(text)
        text = SearchEngine._preprocess_equalize(text)
        return text

    def _mine_tokens(self):
        self.tokens_list = []

        for content in self.content_list:
            text = content
            text = self.preprocess(text)
            tokens = re.split(" ", text)
            self.tokens_list.append(tokens)

    @staticmethod
    def _normalize_remove_postfix(token):
        token = re.sub(r" *تر$", "", token)
        token = re.sub(r" *ترین$", "", token)
        token = re.sub(r" *ها$", "", token)
        token = re.sub(r" *های$", "", token)
        token = re.sub(r" *ات$", "", token)
        token = re.sub(r"(\u200c)+$", "", token)
        return token

    @staticmethod
    def _normalize_remove_prefix(token):
        token = re.sub(r"^می", "", token)
        token = re.sub(r"^(\u200c)+", "", token)
        token = re.sub(r"^\u200f", "", token)
        return token

    def _normalize_lemmatize(self, token):
        for item in self.lemmatize:
            past = item["past"]
            present = item["present"]
            infinitive = past + "ن"

            past_postfix = ["م", "یم", "ی", "ید", "", "ند"]
            present_postfix = ["م", "یم", "ی", "ید", "د", "ند"]

            for postfix in past_postfix:
                verb = past + postfix
                if verb == token:
                    return infinitive

            for postfix in present_postfix:
                verb = present + postfix
                if verb == token:
                    return infinitive

        return token

    def _normalize_mokassar(self, token):
        for item in self.mokassar:
            if token == item["plural"]:
                return item["single"]

        return token

    def normalize(self, token):
        if token not in self.stop_normalization:
            token = self._normalize_remove_prefix(token)
            token = self._normalize_remove_postfix(token)
            token = self._normalize_lemmatize(token)
            token = self._normalize_mokassar(token)
        return token

    def _normalize_tokens(self):
        new_token_list = []
        for tokens in self.tokens_list:
            new_tokens = []
            for token in tokens:
                token = self.normalize(token)
                new_tokens.append(token)
            new_token_list.append(new_tokens)
        self.tokens_list = new_token_list

    def _find_stop_words(self):
        all_words = []
        for tokens in self.tokens_list:
            all_words += tokens

        values, counts = np.unique(all_words, return_counts=True)
        self.stop_words_threshold = 3500
        self.stop_words = []

        for i in range(len(values)):
            if counts[i] > self.stop_words_threshold:
                print(values[i])
                print(counts[i])
                self.stop_words.append(values[i])

        self.stop_words = set(self.stop_words)

    def _calculate_idf(self):
        print(self.inverted_index)

        for item in self.inverted_index:
            df = len(item["docs"])
            idf = log10(self.NUMBER_DOCS / df)
            item["idf"] = idf
            print(item)

    def _calculate_length(self):
        self.length_list = []
        for i in range(self.NUMBER_DOCS):
            self.length_list.append(0)

        for item in self.inverted_index:
            idf = item["idf"]
            for doc in item["docs"]:
                doc_id = doc["id"]
                doc_tf = doc["tf"]
                self.length_list[doc_id - 1] += idf * doc_tf

        for i in range(self.NUMBER_DOCS):
            self.length_list[i] = sqrt(self.length_list[i])

    def _aggregate_inverted_index(self):
        term_doc_list = []
        for i, tokens in enumerate(self.tokens_list):
            unique_tokens_list, unique_tokens_counts = np.unique(tokens, return_counts=True)

            # unique_tokens_list = set(tokens)
            # unique_tokens_list -= self.stop_words
            for j, word in enumerate(unique_tokens_list):
                if word not in self.stop_words:
                    tf = 1 + log10(unique_tokens_counts[j])
                    term_doc_list.append({'term': word, 'doc': i + 1, 'tf': tf})

        sorted_term_doc_list = sorted(term_doc_list, key=lambda x: x["term"])

        self.inverted_index = []
        previous_term = None
        for term_doc in sorted_term_doc_list:
            current_term = term_doc["term"]
            doc_id = term_doc["doc"]
            tf = term_doc["tf"]
            if previous_term != current_term:
                self.inverted_index.append({"term": current_term, "docs": []})
            self.inverted_index[-1]["docs"].append({"id": doc_id, "tf": tf})

            previous_term = current_term

    def _save_inverted_index(self):
        with open('./output_data/inverted_index', 'wb') as fp:
            pickle.dump(self.inverted_index, fp)
        with open('./output_data/dictionary', 'wb') as fp:
            pickle.dump(self.dictionary, fp)

    def load_inverted_index(self):
        self._load_input_data()
        with open('./output_data/inverted_index', 'rb') as fp:
            self.inverted_index = pickle.load(fp)
        with open('./output_data/dictionary', 'rb') as fp:
            self.dictionary = pickle.load(fp)
        self._calculate_length()

    def _get_documents(self, word):
        word = self.preprocess(word)
        word = re.sub(" ", "", word)
        word = self.normalize(word)
        i = bisect_left(self.dictionary, word)
        if i != len(self.dictionary) and word == self.dictionary[i]:
            return self.inverted_index[i]
        else:
            return None

    def _get_context(self, doc_id):
        return self.content_list[doc_id - 1]

    def _get_url(self, doc_id):
        return self.url_list[doc_id - 1]

    def _create_dictionary(self):
        self.dictionary = []
        for x in self.inverted_index:
            self.dictionary.append(x["term"])
            # if len(x["docs"]) > 10:
            print(x["term"])

    def _create_champion_list(self):
        for term_doc in self.inverted_index:
            if len(term_doc["docs"]) <= self.NUMBER_DOCS_CHAMPION_LIST:
                term_doc["champion_docs"] = term_doc["docs"]
            else:
                docs = term_doc["docs"]
                for doc in docs:
                    doc["score"] = 1.0 * doc["tf"] / self.length_list[doc["id"] - 1]

                sorted_doc = sorted(docs, key=lambda x: -x["score"])[:self.NUMBER_DOCS_CHAMPION_LIST]
                for doc in sorted_doc:
                    del doc["score"]
                term_doc["champion_docs"] = sorted(sorted_doc, key=lambda x: x["id"])

    def create_inverted_index(self):
        self._load_input_data()
        self._mine_tokens()
        self._normalize_tokens()
        self._find_stop_words()
        self._aggregate_inverted_index()
        self._create_dictionary()
        self._calculate_idf()
        self._calculate_length()
        self._create_champion_list()
        self._save_inverted_index()

    @staticmethod
    def _is_finish_searching(doc_id_list, pointer_list):
        for i in range(len(doc_id_list)):
            if pointer_list[i] < len(doc_id_list[i]):
                return False
        return True

    @staticmethod
    def _next_pointer(pointer_list, doc_id_list, smallest_doc_id):
        new_pointers = []
        for i, pointer in enumerate(pointer_list):
            if pointer < len(doc_id_list[i]):
                doc_id = doc_id_list[i][pointer]
                if doc_id == smallest_doc_id:
                    new_pointers.append(pointer + 1)
                else:
                    new_pointers.append(pointer)
            else:
                new_pointers.append(pointer)

        return new_pointers

    @staticmethod
    def _get_smallest_doc_id(doc_id_list, pointer_list):
        smallest_doc_id = 1000 * 1000 * 1000
        smallest_doc_id_number = 0
        for i, pointer in enumerate(pointer_list):
            if pointer < len(doc_id_list[i]):
                doc_id = doc_id_list[i][pointer]
                if doc_id < smallest_doc_id:
                    smallest_doc_id = doc_id
                    smallest_doc_id_number = 1
                elif doc_id == smallest_doc_id:
                    smallest_doc_id_number += 1

        return smallest_doc_id, smallest_doc_id_number

    def _search_multi_token(self, tokens):
        start = timeit.default_timer()
        score_dictionary = {}

        unique_tokens, counts = np.unique(tokens, return_counts=True)

        for i, token in enumerate(unique_tokens):
            documents = self._get_documents(token)
            if documents is not None:
                idf = documents["idf"]
                query_tf = 1 + log10(counts[i])
                if self.CHAMPION_LIST_ENABLE:
                    docs = documents["champion_docs"]
                else:
                    docs = documents["docs"]

                for doc in docs:
                    doc_id = doc["id"]
                    doc_tf = doc["tf"]

                    new_score = doc_tf * idf * query_tf
                    if doc_id in score_dictionary:
                        score_dictionary[doc_id] += new_score
                    else:
                        score_dictionary[doc_id] = new_score

        results = []
        if self.HEAP_ENABLE:
            heapify(results)

        for item in score_dictionary:
            result = {"doc_id": item, "URL": self._get_url(item),
                      "score": score_dictionary[item] / self.length_list[item - 1]}

            if self.HEAP_ENABLE:
                heappush(results, (-result["score"], (item, result)))
            else:
                results.append(result)

        final_results = []
        if self.HEAP_ENABLE:
            for i in range(self.NUMBER_RESULTS):
                if i >= len(score_dictionary):
                    break
                heap_item = heappop(results)
                if heap_item is not None:
                    result = heap_item[1][1]
                    final_results.append(result)
        else:
            results = sorted(results, key=lambda x: -x["score"])
            if len(results) < self.NUMBER_RESULTS:
                final_results = results
            else:
                final_results = results[:self.NUMBER_RESULTS]

        stop = timeit.default_timer()

        table = PrettyTable()
        table.field_names = ["Row", "Score", "Doc ID", "URL"]
        for i, result in enumerate(final_results):
            table.add_row([i + 1, result["score"], result["doc_id"], result["URL"]])

        print(table)
        print(f"Time: {(stop - start)*1000} mS ")

    def search(self, query):
        preprocessed_query = self.preprocess(query)
        tokens = preprocessed_query.split(" ")
        self._search_multi_token(tokens)

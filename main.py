from openpyxl import load_workbook
import re
from collections import Counter
import numpy as np

wb = load_workbook("./input_data/IR_Spring2021_ph12_7k.xlsx")
sheet = wb.active

content_list = []
url_list = []
tokens_list = []

for i in range(2, 7002):
    data_id = int(sheet.cell(i, 1).value)
    data_content = sheet.cell(i, 2).value
    data_url = sheet.cell(i, 3).value
    content_list.append(data_content)
    url_list.append(data_url)

for content in content_list:
    text = content
    text = re.sub('\.', ' ', text)
    text = re.sub('،', ' ', text)
    text = re.sub(',', ' ', text)
    text = re.sub('\n', ' ', text)
    text = re.sub(':', ' ', text)
    text = re.sub('؛', ' ', text)
    text = re.sub('"', ' ', text)
    text = re.sub('\(', ' ', text)
    text = re.sub('\)', ' ', text)
    text = re.sub('\]', ' ', text)
    text = re.sub('\[', ' ', text)
    text = re.sub('-', ' ', text)
    text = re.sub('\*', ' ', text)
    text = re.sub('«', ' ', text)
    text = re.sub('»', ' ', text)

    text = re.sub(r' +', ' ', text)

    tokens = re.split(" ", text)
    # sorted_tokens = sorted(tokens)
    tokens_list.append(tokens)


all_words = []
for tokens in tokens_list:
    all_words += tokens


values, counts = np.unique(all_words, return_counts=True)
stop_words_threshold = 1000
stop_words = []

for i in range(len(values)):
    if counts[i] > 1000:
        stop_words.append(values[i])

stop_words = set(stop_words)

term_doc_list = []
for i, tokens_list in enumerate(tokens_list):
    unique_tokens_list = set(tokens_list)
    unique_tokens_list -= stop_words
    for word in unique_tokens_list:
        term_doc_list.append({'term': word, 'doc': i+1})

sorted_term_doc_list = sorted(term_doc_list, key=lambda x: x["term"])

for v in reverse_index:
    print(v)
